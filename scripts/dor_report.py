"""Build the final Distribution-of-Returns xlsx report.

Layout:
- Summary sheet (first): 22 assets x 8 metric columns (C-C StdDev D/W/M/Q,
  H-L Return Avg D/W/M/Q), formatted as percentages.
- Grading sheet (second): same 8 metrics, but each value replaced by a
  quintile grade A-E (computed across the universe in this report) with a
  green-to-red colour scale, for at-a-glance comparison.
- Per-asset DoR sheets (one per asset, per timeframe): OHLC + returns table
  on the left; per-return blocks on the right containing bin distribution,
  descriptive stats, and positive-return analysis.

Inspired by Distribution_of_Returns_Template.xlsx — same information, cleaner
layout (consistent block geometry across timeframes).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PCT_FMT = "0.00%"
PCT4_FMT = "0.0000%"
INT_FMT = "0"

HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
TITLE_FONT = Font(bold=True, size=12)
SUBTITLE_FONT = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True)

GRADE_LETTERS = ["A", "B", "C", "D", "E"]
GRADE_FILLS = [
    PatternFill(start_color="FF63BE7B", end_color="FF63BE7B", fill_type="solid"),  # green
    PatternFill(start_color="FFB6E2A1", end_color="FFB6E2A1", fill_type="solid"),  # light green
    PatternFill(start_color="FFFFEB84", end_color="FFFFEB84", fill_type="solid"),  # yellow
    PatternFill(start_color="FFFCAA7E", end_color="FFFCAA7E", fill_type="solid"),  # light red
    PatternFill(start_color="FFF8696B", end_color="FFF8696B", fill_type="solid"),  # red
]

TF_LABEL = {"d": "Daily", "w": "Weekly", "m": "Monthly", "q": "Quarterly"}
RETURN_COLS_BY_TF = {
    "d": ["C-C Returns", "H-L Returns", "O-C Returns"],
    "w": ["C-C Returns", "H-L Returns"],
    "m": ["C-C Returns", "H-L Returns"],
    "q": ["C-C Returns", "H-L Returns"],
}

STAT_ROWS = [
    ("Mean", "mean", PCT4_FMT),
    ("Standard Error", "standard_error", PCT4_FMT),
    ("Median", "median", PCT4_FMT),
    ("Mode", "mode", PCT4_FMT),
    ("Standard Deviation", "standard_deviation", PCT4_FMT),
    ("Sample Variance", "sample_variance", "0.000000"),
    ("Kurtosis", "kurtosis", "0.0000"),
    ("Skewness", "skewness", "0.0000"),
    ("Range", "range", PCT4_FMT),
    ("Minimum", "minimum", PCT4_FMT),
    ("Maximum", "maximum", PCT4_FMT),
    ("Sum", "sum", PCT4_FMT),
    ("Count", "count", INT_FMT),
]

POSITIVE_ROWS = [
    ("Average Positive", "average_positive", PCT4_FMT),
    ("Count Positive", "count_positive", INT_FMT),
    ("Frequency % Positive", "freq_pct_positive", PCT4_FMT),
    ("Freq-Adjusted Return", "freq_adjusted_return", PCT4_FMT),
]

# Block geometry: each return-block occupies 5 columns plus a 1-col gap.
BLOCK_WIDTH = 5
BLOCK_GAP = 1
BLOCK_STRIDE = BLOCK_WIDTH + BLOCK_GAP


def _safe_sheet_name(name: str) -> str:
    bad = set('[]:*?/\\')
    return "".join("_" if c in bad else c for c in name)[:31]


def _write_summary(ws, asset_results: list[dict]) -> None:
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    for col_idx in range(4, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws["B1"] = "Distribution of Returns — Summary"
    ws["B1"].font = TITLE_FONT
    ws.merge_cells("B1:I1")

    ws["B3"] = "Ticker";       ws["B3"].font = HEADER_FONT
    ws["C3"] = "Asset Class";  ws["C3"].font = HEADER_FONT
    ws.cell(row=2, column=4, value="C-C Standard Deviation").font = HEADER_FONT
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7)
    ws.cell(row=2, column=8, value="H-L Return Average").font = HEADER_FONT
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=11)
    headers = ["Daily", "Weekly", "Monthly", "Quarterly"] * 2
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=4 + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    for r, asset in enumerate(asset_results, start=4):
        ws.cell(row=r, column=2, value=asset["ticker"])
        ws.cell(row=r, column=3, value=asset["asset_class"])
        for tf_idx, tf in enumerate(("d", "w", "m", "q")):
            stats = asset["by_tf"][tf].get("C-C Returns", {}).get("stats", {})
            cell = ws.cell(row=r, column=4 + tf_idx, value=stats.get("standard_deviation"))
            cell.number_format = PCT_FMT
        for tf_idx, tf in enumerate(("d", "w", "m", "q")):
            stats = asset["by_tf"][tf].get("H-L Returns", {}).get("stats", {})
            cell = ws.cell(row=r, column=8 + tf_idx, value=stats.get("mean"))
            cell.number_format = PCT_FMT


def _quintile_grades(values: list) -> list:
    """Rank values into 5 buckets [0..4]; lower value -> 0 (best/least). None passes through."""
    indexed = [(v, i) for i, v in enumerate(values) if v is not None]
    n = len(indexed)
    grades: list = [None] * len(values)
    if n == 0:
        return grades
    indexed.sort(key=lambda x: x[0])
    for rank, (_, original_idx) in enumerate(indexed):
        grades[original_idx] = min(int(rank * 5 / n), 4)
    return grades


def _write_grading(ws, asset_results: list[dict]) -> None:
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    for col_idx in range(4, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    ws["B1"] = "Distribution of Returns — Grading (quintile ranks across this universe)"
    ws["B1"].font = TITLE_FONT
    ws.merge_cells("B1:K1")

    ws["B3"] = "Ticker";       ws["B3"].font = HEADER_FONT
    ws["C3"] = "Asset Class";  ws["C3"].font = HEADER_FONT
    ws.cell(row=2, column=4, value="C-C Standard Deviation").font = HEADER_FONT
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7)
    ws.cell(row=2, column=8, value="H-L Return Average").font = HEADER_FONT
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=11)
    headers = ["Daily", "Weekly", "Monthly", "Quarterly"] * 2
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=4 + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    cc_vals = {tf: [] for tf in ("d", "w", "m", "q")}
    hl_vals = {tf: [] for tf in ("d", "w", "m", "q")}
    for asset in asset_results:
        for tf in ("d", "w", "m", "q"):
            cc_vals[tf].append(asset["by_tf"][tf].get("C-C Returns", {}).get("stats", {}).get("standard_deviation"))
            hl_vals[tf].append(asset["by_tf"][tf].get("H-L Returns", {}).get("stats", {}).get("mean"))
    cc_grades = {tf: _quintile_grades(cc_vals[tf]) for tf in ("d", "w", "m", "q")}
    hl_grades = {tf: _quintile_grades(hl_vals[tf]) for tf in ("d", "w", "m", "q")}

    for r, asset in enumerate(asset_results):
        ws.cell(row=4 + r, column=2, value=asset["ticker"])
        ws.cell(row=4 + r, column=3, value=asset["asset_class"])
        for tf_idx, tf in enumerate(("d", "w", "m", "q")):
            for col_offset, grades, vals in (
                (4, cc_grades[tf], cc_vals[tf]),
                (8, hl_grades[tf], hl_vals[tf]),
            ):
                grade = grades[r]
                cell = ws.cell(row=4 + r, column=col_offset + tf_idx)
                if grade is None:
                    continue
                cell.value = f"{GRADE_LETTERS[grade]}  ({vals[r]:.2%})"
                cell.fill = GRADE_FILLS[grade]
                cell.alignment = Alignment(horizontal="center")
                cell.font = HEADER_FONT

    legend_row = 4 + len(asset_results) + 2
    ws.cell(row=legend_row, column=2, value="Legend").font = SUBTITLE_FONT
    descriptions = [
        ("A", "Lowest 20% — least volatile / tightest range"),
        ("B", "20–40%"),
        ("C", "40–60% — median band"),
        ("D", "60–80%"),
        ("E", "Highest 20% — most volatile / widest range"),
    ]
    for i, (letter, desc) in enumerate(descriptions):
        cell = ws.cell(row=legend_row + 1 + i, column=2, value=letter)
        cell.fill = GRADE_FILLS[i]
        cell.alignment = Alignment(horizontal="center")
        cell.font = HEADER_FONT
        ws.cell(row=legend_row + 1 + i, column=3, value=desc)
    ws.cell(
        row=legend_row + len(descriptions) + 2, column=2,
        value="Ranks are recomputed per column (metric × timeframe) using the assets present in this report.",
    )


def _write_ohlc_block(ws, frame: pd.DataFrame, tf: str) -> int:
    """Write OHLC + returns table starting at A3. Returns last row used."""
    columns = ["Open", "High", "Low", "Close", "Adj Close"] + RETURN_COLS_BY_TF[tf]
    columns = [c for c in columns if c in frame.columns]
    headers = ["Date"] + columns
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    # Cap rows actually written to keep the file small when daily history is huge.
    MAX_ROWS = 5000
    head = frame.head(MAX_ROWS)
    for r_offset, (date_str, row) in enumerate(head.iterrows(), start=4):
        ws.cell(row=r_offset, column=1, value=date_str)
        for col_idx, col in enumerate(columns, start=2):
            val = row[col]
            cell = ws.cell(
                row=r_offset, column=col_idx,
                value=None if pd.isna(val) else float(val),
            )
            if col in ("C-C Returns", "H-L Returns", "O-C Returns"):
                cell.number_format = PCT4_FMT
            else:
                cell.number_format = "0.0000"
    if len(frame) > MAX_ROWS:
        ws.cell(row=4 + MAX_ROWS, column=1, value=f"... ({len(frame) - MAX_ROWS} earlier rows truncated for readability)")
    # Column widths
    ws.column_dimensions["A"].width = 11
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    return 4 + min(len(head), MAX_ROWS)


def _write_return_block(ws, start_col: int, return_col: str, payload: dict) -> None:
    """Write one return-column block (bins + stats + positives) starting at start_col, row 3."""
    bins: pd.DataFrame = payload["bins"]
    stats: dict = payload["stats"]
    positives: dict = payload["positives"]

    title_cell = ws.cell(row=3, column=start_col, value=return_col)
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col + BLOCK_WIDTH - 1)

    # Bins table
    row = 5
    ws.cell(row=row, column=start_col, value="Distribution").font = SUBTITLE_FONT
    row += 1
    headers = ["Bin", "Edge", "Frequency", "Probability", "Cumulative"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    bins_start = row
    for _, br in bins.iterrows():
        ws.cell(row=row, column=start_col + 0, value=br["k_label"])
        edge_cell = ws.cell(row=row, column=start_col + 1, value=None if pd.isna(br["edge"]) else float(br["edge"]))
        edge_cell.number_format = PCT4_FMT
        ws.cell(row=row, column=start_col + 2, value=int(br["frequency"])).number_format = INT_FMT
        ws.cell(row=row, column=start_col + 3, value=float(br["probability"])).number_format = PCT_FMT
        ws.cell(row=row, column=start_col + 4, value=float(br["cumulative"])).number_format = PCT_FMT
        row += 1

    row += 1  # gap

    # Descriptive stats
    ws.cell(row=row, column=start_col, value="Descriptive Statistics").font = SUBTITLE_FONT
    row += 1
    for label, key, fmt in STAT_ROWS:
        ws.cell(row=row, column=start_col, value=label)
        cell = ws.cell(row=row, column=start_col + 1, value=stats.get(key))
        cell.number_format = fmt
        row += 1

    row += 1  # gap

    # Positive return analysis
    ws.cell(row=row, column=start_col, value="Positive Return Analysis").font = SUBTITLE_FONT
    row += 1
    for label, key, fmt in POSITIVE_ROWS:
        ws.cell(row=row, column=start_col, value=label)
        cell = ws.cell(row=row, column=start_col + 1, value=positives.get(key))
        cell.number_format = fmt
        row += 1

    # Block column widths
    ws.column_dimensions[get_column_letter(start_col)].width = 22
    for off in range(1, BLOCK_WIDTH):
        ws.column_dimensions[get_column_letter(start_col + off)].width = 13


def _write_dor_sheet(wb: Workbook, asset: dict, tf: str) -> None:
    sheet_name = _safe_sheet_name(f"{asset['ticker']}_{TF_LABEL[tf]}")
    ws = wb.create_sheet(title=sheet_name)
    ws["A1"] = f"{asset['display_name']} ({asset['ticker']}) — {TF_LABEL[tf]} Distribution of Returns"
    ws["A1"].font = TITLE_FONT

    frame = asset["frames"][tf]
    _write_ohlc_block(ws, frame, tf)

    # Place return blocks starting at column K (col index 11).
    start_col = 11
    for i, return_col in enumerate(RETURN_COLS_BY_TF[tf]):
        if return_col not in asset["by_tf"][tf]:
            continue
        _write_return_block(
            ws, start_col + i * BLOCK_STRIDE, return_col, asset["by_tf"][tf][return_col]
        )

    ws.freeze_panes = "A4"


def write_workbook(asset_results: list[dict], out_path: Path) -> Path:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _write_summary(summary, asset_results)

    grading = wb.create_sheet("Grading", 1)
    _write_grading(grading, asset_results)

    for asset in asset_results:
        for tf in ("d", "w", "m", "q"):
            _write_dor_sheet(wb, asset, tf)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
