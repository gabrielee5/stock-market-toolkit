"""Compare Python descriptive stats against the Excel template cell-by-cell.

Reads OHLC from each Template *.* DoR sheet, runs the Python pipeline on it,
then reads the Excel-computed stats block and reports any deltas above the
configured tolerance. Validates that the Python implementation matches Excel's
formulas to floating-point precision on identical inputs.

Usage:
    python scripts/verify_against_excel.py [TEMPLATE_PATH]
"""

from __future__ import annotations

import math
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from dor import descriptive_stats

DEFAULT_TEMPLATE = Path("/Users/gabrielefabietti/Desktop/IPLT/iplt-video4-resources/Distribution_of_Returns_Template.xlsx")
REL_TOL = 1e-9
ABS_TOL = 1e-12

# Per-sheet layout: where the OHLC data lives + which stat block sits where.
# block format: (return_col, label_col_letter, value_col_letter, start_row)
SHEETS = {
    "Template Daily DoR": {
        "blocks": [
            ("C-C Returns", "K", "L", 17),
            ("H-L Returns", "R", "S", 17),
            ("O-C Returns", "Y", "Z", 17),
        ],
    },
    "Template Weekly DoR": {
        "blocks": [
            ("C-C Returns", "J", "K", 17),
            ("H-L Returns", "Q", "R", 17),
        ],
    },
    "Template Monthly DoR": {
        "blocks": [
            ("C-C Returns", "J", "K", 19),
            ("H-L Returns", "Q", "R", 19),
        ],
    },
    "Template Quarterly DoR": {
        "blocks": [
            ("C-C Returns", "J", "K", 19),
            ("H-L Returns", "Q", "R", 19),
        ],
    },
}

# Maps the Excel stat label to the descriptive_stats() key.
STAT_LABELS = {
    "Mean":               "mean",
    "Standard Error":     "standard_error",
    "Median":             "median",
    "Mode":               "mode",
    "Standard Deviation": "standard_deviation",
    "Sample Variance":    "sample_variance",
    "Kurtosis":           "kurtosis",
    "Skewness":           "skewness",
    "Range":              "range",
    "Minimum":            "minimum",
    "Maximum":            "maximum",
    "Sum":                "sum",
    "Count":              "count",
}


def _col_idx(letter: str) -> int:
    return ord(letter.upper()) - ord("A") + 1


def load_ohlc_from_sheet(ws) -> pd.DataFrame:
    """Read OHLC + Adj Close from columns A:F starting at row 2.

    The template stores newest-first (matching the project convention).
    Returns a DataFrame indexed by Date string 'DD-MM-YY'.
    """
    rows = []
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        o = ws.cell(row=r, column=2).value
        if d is None or o is None:
            continue
        rows.append({
            "Date": pd.to_datetime(d).strftime("%d-%m-%y"),
            "Open": float(o),
            "High": float(ws.cell(row=r, column=3).value),
            "Low":  float(ws.cell(row=r, column=4).value),
            "Close": float(ws.cell(row=r, column=5).value),
            "Adj Close": float(ws.cell(row=r, column=6).value),
        })
    return pd.DataFrame(rows).set_index("Date")


def compute_returns(frame: pd.DataFrame, has_oc: bool) -> pd.DataFrame:
    """Compute C-C / H-L / O-C returns the same way dor.clean_data does."""
    out = frame.copy()
    prev_adj = out["Adj Close"].shift(-1)
    out["C-C Returns"] = (out["Adj Close"] - prev_adj) / prev_adj
    out["H-L Returns"] = (out["High"] - out["Low"]) / out["Low"]
    if has_oc:
        out["O-C Returns"] = (out["Close"] - out["Open"]) / out["Open"]
    # Don't drop the oldest row globally; C-C is NaN there but H-L (and O-C) are
    # valid. Excel includes that row in COUNT(H), so we must too. descriptive_stats
    # handles NaN per-column.
    return out


def excel_stats_block(ws_values, label_col: str, value_col: str, start_row: int) -> dict:
    """Read the 13-row descriptive-stats block and map labels to numeric values."""
    out = {}
    for offset in range(13):
        row = start_row + offset
        label = ws_values.cell(row=row, column=_col_idx(label_col)).value
        if not label:
            continue
        key = STAT_LABELS.get(str(label).strip())
        if key is None:
            continue
        out[key] = ws_values.cell(row=row, column=_col_idx(value_col)).value
    return out


def compare(label: str, py_val, xl_val, tol_rel=REL_TOL, tol_abs=ABS_TOL) -> tuple[bool, str]:
    if xl_val is None or (isinstance(xl_val, str) and xl_val.startswith("#")):
        # Excel #N/A (e.g. MODE on continuous data); Python should be None.
        ok = py_val is None
        return ok, f"  {label:22s} excel={xl_val!r:25s} python={py_val!r:15s} -> {'OK' if ok else 'MISMATCH'}"
    if py_val is None:
        return False, f"  {label:22s} excel={xl_val!r:25s} python=None -> MISMATCH"
    ok = math.isclose(float(py_val), float(xl_val), rel_tol=tol_rel, abs_tol=tol_abs)
    delta = abs(float(py_val) - float(xl_val))
    return ok, f"  {label:22s} excel={float(xl_val):.10g} python={float(py_val):.10g} delta={delta:.2e} -> {'OK' if ok else 'MISMATCH'}"


def main(template_path: Path) -> int:
    print(f"Verifying against {template_path}\n")
    wb_formulas = load_workbook(template_path, data_only=False)
    wb_values = load_workbook(template_path, data_only=True)
    total_ok = 0
    total_mismatch = 0

    for sheet_name, layout in SHEETS.items():
        print(f"=== {sheet_name} ===")
        ws_v = wb_values[sheet_name]
        frame = load_ohlc_from_sheet(ws_v)
        if frame.empty:
            print(f"  (no OHLC data — skipping)\n")
            continue
        has_oc = "Daily" in sheet_name
        with_returns = compute_returns(frame, has_oc=has_oc)
        for return_col, label_col, value_col, start_row in layout["blocks"]:
            if return_col not in with_returns.columns:
                continue
            print(f"  -- {return_col} --")
            py_stats = descriptive_stats(with_returns[return_col])
            xl_stats = excel_stats_block(ws_v, label_col, value_col, start_row)
            for label, key in STAT_LABELS.items():
                if key not in xl_stats:
                    continue
                ok, msg = compare(label, py_stats.get(key), xl_stats[key])
                print(msg)
                if ok:
                    total_ok += 1
                else:
                    total_mismatch += 1
        print()

    print(f"Total: {total_ok} OK, {total_mismatch} MISMATCH")
    return 0 if total_mismatch == 0 else 1


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_TEMPLATE
    sys.exit(main(path))
